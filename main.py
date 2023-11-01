import customtkinter
from tkinter import *
from tkinter import StringVar
import openpyxl

customtkinter.set_appearance_mode("dark")
customtkinter.set_default_color_theme("dark-blue")

main = customtkinter.CTk()
main.geometry("600x500")
main.resizable(False, False)
main.title("PLANILHA ATENDIMENTOS")

font = customtkinter.CTkFont(family="Times", 
                             size=15, 
                             weight="bold", 
                             slant="italic"
                        ) 

dataGet = ""
localGet = ""
solicGet = ""
probGet = ""
concGet = ""
initGet = ""
finalGet = ""

data_var = StringVar()
local_var = StringVar()
solic_var = StringVar() 
prob_var = StringVar()
conc_var = StringVar()
init_var = StringVar()
final_var = StringVar()

def info():
    global data_var, local_var, solic_var, prob_var, conc_var, init_var, final_var

    dataGet = data_var.get()
    localGet = local_var.get()
    solicGet = solic_var.get()
    probGet = prob_var.get()
    concGet = conc_var.get()
    initGet = init_var.get()

    data_var.set("")
    local_var.set("")
    solic_var.set("")
    prob_var.set("")
    conc_var.set("")
    init_var.set("")

    plan = openpyxl.load_workbook('atend.xlsx')

    for sheet_name in plan.sheetnames:
        sheet = plan[sheet_name]

        rowNumb = 1

        while sheet.cell(row=rowNumb, column=1).value is not None:
            rowNumb += 1

        data = dataGet
        local = localGet
        solicitante = solicGet
        problema = probGet
        conclusao = concGet
        init = initGet

        sheet.cell(row=rowNumb, column=1, value=data)
        sheet.cell(row=rowNumb, column=2, value=local)
        sheet.cell(row=rowNumb, column=3, value=solicitante)
        sheet.cell(row=rowNumb, column=4, value=problema)
        sheet.cell(row=rowNumb, column=5, value=conclusao)
        sheet.cell(row=rowNumb, column=6, value=init)

        
        plan.save('atend.xlsx')




frame = customtkinter.CTkFrame(master = main, width=600, height=500, fg_color="#221d24")
frame.pack()

CPD = customtkinter.CTkLabel(master=frame, text="𝘾𝙚𝙣𝙩𝙧𝙤 𝙙𝙚 𝙋𝙧𝙤𝙘𝙚𝙨𝙨𝙖𝙢𝙚𝙣𝙩𝙤 𝙙𝙚 𝘿𝙖𝙙𝙤𝙨")
CPD.place(x=295, y=190)

labelData = customtkinter.CTkLabel(master=frame, text="✏️ Data", font=font)
labelData.place(x=10, y=50)
data = customtkinter.CTkEntry(master=frame, width=100, textvariable=data_var)
data.place(x=10, y=80)

labelLocal = customtkinter.CTkLabel(master=frame, text="☢ Local", font=font)
labelLocal.place(x=10, y=120)
local = customtkinter.CTkEntry(master=frame, width=200, textvariable=local_var)
local.place(x=10, y=150)

labelSolicit = customtkinter.CTkLabel(master=frame, text="☎ Solicitante", font=font)
labelSolicit.place(x=10, y=190)
localSolicit = customtkinter.CTkEntry(master=frame, width=200,textvariable=solic_var)
localSolicit.place(x=10, y=220)

labelProb = customtkinter.CTkLabel(master=frame, text="➮ Problema", font=font)
labelProb.place(x=10, y=260)
localProb = customtkinter.CTkEntry(master=frame, width=200, textvariable=prob_var)
localProb.place(x=10, y=290)

labelConc = customtkinter.CTkLabel(master=frame, text="✓ Conclusão", font=font)
labelConc.place(x=10, y=330)
localConc = customtkinter.CTkEntry(master=frame, width=200, textvariable=conc_var)
localConc.place(x=10, y=360)

labelinit = customtkinter.CTkLabel(master=frame, text="⌛  Horário de inicio e término", font=font)
labelinit.place(x=10, y=400)
localinit = customtkinter.CTkEntry(master=frame, width=200, textvariable=init_var)
localinit.place(x=10, y=430)


but = customtkinter.CTkButton(master=frame, text="preencher", width=200, command=info).place(x=298, y=260)

img = PhotoImage(file="logorpef.png")
imgLabel = Label(image=img, bg="#221d24")
imgLabel.place(x=303, y=30)

main.mainloop()
